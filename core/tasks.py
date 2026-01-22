from celery import shared_task
from openpyxl import load_workbook
from pathlib import Path
from .models import Customer, Loan


DATA_DIR = Path(__file__).resolve().parent.parent / 'data'


@shared_task
def ingest_customers():
    """Load customer data from Excel into DB."""
    wb = load_workbook(DATA_DIR / 'customer_data.xlsx')
    ws = wb.active
    
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        Customer.objects.update_or_create(
            id=row[0],
            defaults={
                'first_name': row[1] or '',
                'last_name': row[2] or '',
                'age': row[3],
                'phone_number': row[4] or 0,
                'monthly_salary': row[5] or 0,
                'approved_limit': row[6] or 0,
            }
        )
        count += 1
    
    return f'Imported {count} customers'


@shared_task
def ingest_loans():
    wb = load_workbook(DATA_DIR / 'loan_data.xlsx')
    ws = wb.active
    
    imported = 0
    skipped = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        
        # Skip if customer doesn't exist
        cust = Customer.objects.filter(id=row[0]).first()
        if not cust:
            skipped += 1
            continue
        
        Loan.objects.update_or_create(
            id=row[1],
            defaults={
                'customer': cust,
                'loan_amount': row[2] or 0,
                'tenure': row[3] or 0,
                'interest_rate': row[4] or 0,
                'monthly_repayment': row[5] or 0,
                'emis_paid_on_time': row[6] or 0,
                'start_date': row[7],
                'end_date': row[8],
            }
        )
        imported += 1
    
    return f'Imported {imported} loans, skipped {skipped} (missing customer)'
